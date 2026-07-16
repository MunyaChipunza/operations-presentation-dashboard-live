from __future__ import annotations

import argparse
import base64
import csv
import datetime as dt
import html
import json
import os
import re
import shutil
import statistics
import subprocess
import tempfile
import time
import urllib.parse
import urllib.request
import zipfile
from pathlib import Path
from typing import Any, Callable

from openpyxl import Workbook, load_workbook


WORKBOOK_NAME_HINTS = ("PPT presentation source data", "Operations Data")
CSV_SOURCE_HINTS = ("Operations Data", "PPT presentation source data")
TIMEZONE_NAME = "Africa/Johannesburg"
DEFAULT_OUTPUT = "dashboard_data.json"
LIVE_DATA_MAX_ROW = 100
CREATE_NO_WINDOW = getattr(subprocess, "CREATE_NO_WINDOW", 0)
EXCEL_SUFFIXES = {".xlsx", ".xlsm"}
CSV_SUFFIX = ".csv"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build dashboard JSON from the PPT presentation source workbook.")
    parser.add_argument("--workbook", help="Local workbook or CSV export path (.xlsx/.xlsm/.csv).")
    parser.add_argument("--workbook-url", help="Public share or direct download URL for the workbook.")
    parser.add_argument("--output", default=DEFAULT_OUTPUT, help="Where to write the dashboard JSON.")
    return parser.parse_args()


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def to_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    text = clean_text(value).replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def to_int(value: Any) -> int | None:
    number = to_float(value)
    if number is None:
        return None
    return int(round(number))


def is_whole_number(value: float) -> bool:
    return abs(value - round(value)) < 1e-9


def format_number(value: float | None, digits: int = 0) -> str:
    if value is None:
        return "-"
    if digits == 0 and is_whole_number(value):
        return f"{int(round(value)):,}"
    return f"{value:,.{digits}f}"


def format_percent(value: float | None, digits: int = 0) -> str:
    if value is None:
        return "-"
    return f"{value * 100:.{digits}f}%"


def format_month_label(value: Any) -> str:
    text = clean_text(value)
    if not text:
        return ""
    normalized = text[:3].title()
    if normalized == "Sep":
        return "Sep"
    return normalized


def mean(values: list[float]) -> float | None:
    if not values:
        return None
    return statistics.fmean(values)


def path_is_csv(path: Path) -> bool:
    return path.suffix.lower() == CSV_SUFFIX


def path_is_excel(path: Path) -> bool:
    return path.suffix.lower() in EXCEL_SUFFIXES


def safe_key(text: str, fallback: str) -> str:
    cleaned = re.sub(r"[^0-9a-zA-Z]+", "_", text.strip().lower()).strip("_")
    if not cleaned:
        return fallback
    if cleaned[0].isdigit():
        cleaned = f"n_{cleaned}"
    return cleaned


def candidate_source_dirs(bundle_dir: Path) -> list[Path]:
    roots: list[Path] = []
    for candidate in (Path.home() / "Downloads", bundle_dir.parent, bundle_dir.parent.parent):
        if candidate.exists() and candidate not in roots:
            roots.append(candidate)
    return roots


def iter_matching_files(roots: list[Path], patterns: tuple[str, ...]) -> list[Path]:
    matches: list[Path] = []
    seen: set[Path] = set()
    for root in roots:
        for pattern in patterns:
            for candidate in root.glob(pattern):
                if candidate.is_file() and candidate not in seen:
                    seen.add(candidate)
                    matches.append(candidate)
    return matches


def latest_path(paths: list[Path]) -> Path | None:
    if not paths:
        return None
    return max(paths, key=lambda candidate: candidate.stat().st_mtime)


def find_default_csv_export(bundle_dir: Path) -> Path | None:
    roots = candidate_source_dirs(bundle_dir)
    patterns = tuple(f"*{hint}*.csv" for hint in CSV_SOURCE_HINTS)
    return latest_path(iter_matching_files(roots, patterns))


def choose_preferred_source(candidate_path: Path | None, bundle_dir: Path) -> Path | None:
    workbook_path = candidate_path.resolve() if candidate_path and candidate_path.exists() else None
    if workbook_path and path_is_csv(workbook_path):
        return workbook_path

    newer_csv = find_default_csv_export(bundle_dir)
    if newer_csv and (workbook_path is None or newer_csv.stat().st_mtime > workbook_path.stat().st_mtime):
        return newer_csv.resolve()

    if workbook_path:
        return workbook_path

    return find_default_workbook(bundle_dir)


def tone_from_percent(value: float | None, *, good: float, warn: float, higher_is_better: bool = True) -> str:
    if value is None:
        return "quiet"
    if higher_is_better:
        if value >= good:
            return "good"
        if value >= warn:
            return "warn"
        return "bad"
    if value <= good:
        return "good"
    if value <= warn:
        return "warn"
    return "bad"


def tone_housekeeping(value: float | None) -> str:
    if value is None:
        return "quiet"
    if value < 0.50:
        return "bad"
    if value < 0.80:
        return "warn"
    return "good"


def tone_accuracy(value: float | None) -> str:
    if value is None:
        return "quiet"
    if value <= 0.97:
        return "bad"
    if value < 0.99:
        return "warn"
    return "good"


def tone_urgent_orders(value: float | None) -> str:
    if value is None:
        return "quiet"
    if value <= 0.05:
        return "good"
    if value <= 0.10:
        return "warn"
    return "bad"


def tone_from_rank(rank: int | None) -> str:
    if rank is None:
        return "quiet"
    if rank == 1:
        return "good"
    if rank == 2:
        return "warn"
    return "bad"


def tone_assembly_fill_rate(value: float | None) -> str:
    if value is None:
        return "quiet"
    if value < 0.70:
        return "bad"
    if value < 0.85:
        return "warn"
    return "good"


def latest_non_null(rows: list[dict[str, Any]], key: str) -> dict[str, Any] | None:
    for row in reversed(rows):
        if row.get(key) is not None:
            return row
    return None


def max_row(rows: list[dict[str, Any]], key: str) -> dict[str, Any] | None:
    with_values = [row for row in rows if row.get(key) is not None]
    if not with_values:
        return None
    return max(with_values, key=lambda row: row[key])


def min_row(rows: list[dict[str, Any]], key: str) -> dict[str, Any] | None:
    with_values = [row for row in rows if row.get(key) is not None]
    if not with_values:
        return None
    return min(with_values, key=lambda row: row[key])


def build_trendline(values: list[float | None]) -> list[float | None]:
    points = [(index, value) for index, value in enumerate(values) if value is not None]
    if len(points) < 2:
        return [None for _ in values]

    x_values = [point[0] for point in points]
    y_values = [point[1] for point in points]
    x_mean = statistics.fmean(x_values)
    y_mean = statistics.fmean(y_values)
    denominator = sum((x - x_mean) ** 2 for x in x_values)
    slope = 0 if denominator == 0 else sum((x - x_mean) * (y - y_mean) for x, y in points) / denominator
    intercept = y_mean - slope * x_mean
    return [intercept + slope * index for index in range(len(values))]


def constant_series(length: int, value: float | None) -> list[float | None]:
    if value is None:
        return [None for _ in range(length)]
    return [value for _ in range(length)]


def add_target_line(
    dataset: dict[str, Any],
    *,
    value: float | None,
    label: str,
    format_type: str,
    color: str = "#39ff88",
) -> None:
    if value is None:
        return
    labels = dataset["chart"].get("labels", [])
    dataset["chart"]["series"].append(
        {
            "name": label,
            "format": format_type,
            "color": color,
            "values": constant_series(len(labels), value),
            "style": "dashed",
            "showDots": False,
            "strokeWidth": 2,
        }
    )


def add_trendline(
    dataset: dict[str, Any],
    *,
    source_key: str,
    label: str,
    color: str = "#f8fafc",
) -> None:
    source = next((item for item in dataset["chart"]["series"] if item.get("key") == source_key), None)
    if not source:
        return
    dataset["chart"]["series"].append(
        {
            "name": label,
            "format": source["format"],
            "color": color,
            "values": build_trendline(source["values"]),
            "style": "dotted",
            "showDots": False,
            "strokeWidth": 2,
        }
    )


def share_token(url: str) -> str:
    raw = base64.b64encode(url.encode("utf-8")).decode("ascii").rstrip("=")
    return "u!" + raw.replace("/", "_").replace("+", "-")


def request_json(url: str, headers: dict[str, str] | None = None, data: bytes | None = None) -> dict[str, Any]:
    request = urllib.request.Request(url, headers=headers or {}, data=data)
    with urllib.request.urlopen(request, timeout=60) as response:
        return json.loads(response.read().decode("utf-8"))


def request_bytes(url: str, headers: dict[str, str] | None = None) -> bytes:
    request = urllib.request.Request(url, headers=headers or {})
    with urllib.request.urlopen(request, timeout=60) as response:
        return response.read()


def write_temp_workbook(payload: bytes, suffix: str) -> Path:
    fd, temp_path = tempfile.mkstemp(suffix=suffix)
    os.close(fd)
    path = Path(temp_path)
    path.write_bytes(payload)
    return path


def coerce_csv_cell(value: Any) -> Any:
    text = clean_text(value)
    if not text:
        return None

    if text.endswith("%"):
        percent_value = to_float(text[:-1])
        if percent_value is not None:
            return percent_value / 100

    numeric_text = text.replace(",", "")
    if re.fullmatch(r"[-+]?\d*\.?\d+", numeric_text):
        numeric_value = float(numeric_text)
        if is_whole_number(numeric_value):
            return int(round(numeric_value))
        return numeric_value

    return text


def read_csv_rows(csv_path: Path) -> list[list[str]]:
    with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.reader(handle))


def csv_value(rows: list[list[str]], row_num: int, col_num: int) -> Any:
    row_index = row_num - 1
    col_index = col_num - 1
    if row_index < 0 or row_index >= len(rows):
        return None
    row = rows[row_index]
    if col_index < 0 or col_index >= len(row):
        return None
    return coerce_csv_cell(row[col_index])


def raw_csv_value(rows: list[list[str]], row_num: int, col_num: int) -> str:
    row_index = row_num - 1
    col_index = col_num - 1
    if row_index < 0 or row_index >= len(rows):
        return ""
    row = rows[row_index]
    if col_index < 0 or col_index >= len(row):
        return ""
    return clean_text(row[col_index])


def find_csv_header_column(rows: list[list[str]], header: str) -> int | None:
    if len(rows) < 2:
        return None
    header_row = rows[1]
    for index, value in enumerate(header_row, start=1):
        if clean_text(value) == header:
            return index
    return None


def create_csv_snapshot(csv_path: Path) -> Path:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "DATA"
    rows = read_csv_rows(csv_path)
    for row_num, row in enumerate(rows, start=1):
        for col_num, value in enumerate(row, start=1):
            ws.cell(row=row_num, column=col_num).value = coerce_csv_cell(value)

    fd, temp_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    output_path = Path(temp_path)
    workbook.save(output_path)
    workbook.close()
    return output_path


def clear_sheet_block(ws, *, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    for row_num in range(start_row, end_row + 1):
        for col_num in range(start_col, end_col + 1):
            ws.cell(row=row_num, column=col_num).value = None


def set_fill_rate_formula(ws, row_num: int) -> None:
    ws.cell(row=row_num, column=33).value = f'=IF(OR(AE{row_num}="",AF{row_num}="",AE{row_num}=0),"",1-AF{row_num}/AE{row_num})'


def sync_local_workbook_from_csv(csv_path: Path, workbook_path: Path) -> bool:
    if not path_is_csv(csv_path) or not path_is_excel(workbook_path):
        return False

    rows = read_csv_rows(csv_path)
    if not rows:
        return False

    workbook = load_workbook(workbook_path)
    try:
        ws = workbook["DATA"]
        workbook_has_george_split = clean_text(ws.cell(row=2, column=28).value).startswith("2026")
        csv_secondary_header = raw_csv_value(rows, 2, 28)
        csv_has_split_2026 = csv_secondary_header.startswith("2026")
        existing_george_values = {row_num: to_float(ws.cell(row=row_num, column=28).value) for row_num in range(3, LIVE_DATA_MAX_ROW + 1)}
        csv_month_col = find_csv_header_column(rows, "Month")
        csv_assembled_col = find_csv_header_column(rows, "Assembled")
        csv_backorders_col = find_csv_header_column(rows, "Backorders")
        csv_fill_rate_col = find_csv_header_column(rows, "Fill Rate")
        if not all((csv_month_col, csv_assembled_col, csv_backorders_col, csv_fill_rate_col)):
            raise ValueError("Could not find the Liseo assembly headers in the CSV export.")

        clear_sheet_block(ws, start_row=3, end_row=LIVE_DATA_MAX_ROW, start_col=1, end_col=27)
        clear_sheet_block(ws, start_row=3, end_row=LIVE_DATA_MAX_ROW, start_col=29, end_col=33)

        for row_num in range(3, LIVE_DATA_MAX_ROW + 1):
            for col_num in range(1, 28):
                ws.cell(row=row_num, column=col_num).value = csv_value(rows, row_num, col_num)

            if workbook_has_george_split:
                total_2026 = to_float(ws.cell(row=row_num, column=27).value)
                if csv_has_split_2026:
                    ws.cell(row=row_num, column=28).value = csv_value(rows, row_num, 28)
                else:
                    preserved_george = existing_george_values.get(row_num)
                    if total_2026 is None:
                        ws.cell(row=row_num, column=28).value = preserved_george
                    elif preserved_george is None:
                        ws.cell(row=row_num, column=28).value = None
                    else:
                        adjusted_cpt = total_2026 - preserved_george
                        if adjusted_cpt < 0:
                            ws.cell(row=row_num, column=28).value = None
                        else:
                            ws.cell(row=row_num, column=27).value = adjusted_cpt
                            ws.cell(row=row_num, column=28).value = preserved_george

            ws.cell(row=row_num, column=29).value = None
            ws.cell(row=row_num, column=30).value = csv_value(rows, row_num, csv_month_col)
            ws.cell(row=row_num, column=31).value = csv_value(rows, row_num, csv_assembled_col)
            ws.cell(row=row_num, column=32).value = csv_value(rows, row_num, csv_backorders_col)

            csv_fill_rate_value = csv_value(rows, row_num, csv_fill_rate_col)
            if csv_fill_rate_value is not None:
                ws.cell(row=row_num, column=33).value = csv_fill_rate_value
            elif csv_value(rows, row_num, csv_assembled_col) is not None or csv_value(rows, row_num, csv_backorders_col) is not None:
                set_fill_rate_formula(ws, row_num)
            else:
                ws.cell(row=row_num, column=33).value = None

        calculation = getattr(workbook, "calculation", None)
        if calculation is not None:
            calculation.fullCalcOnLoad = True
            calculation.forceFullCalc = True

        workbook.save(workbook_path)
        return True
    finally:
        workbook.close()


def create_excel_snapshot(workbook_path: Path) -> Path:
    helper_script = Path(__file__).with_name("save_excel_snapshot.ps1")
    if not helper_script.exists():
        raise FileNotFoundError(f"Snapshot helper not found: {helper_script}")

    fd, temp_path = tempfile.mkstemp(suffix=workbook_path.suffix or ".xlsx")
    os.close(fd)
    snapshot_path = Path(temp_path)

    try:
        shutil.copy2(workbook_path, snapshot_path)
        return snapshot_path
    except PermissionError:
        snapshot_path.unlink(missing_ok=True)
    except OSError:
        snapshot_path.unlink(missing_ok=True)

    startupinfo = None
    if os.name == "nt":
        startupinfo = subprocess.STARTUPINFO()
        startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
        startupinfo.wShowWindow = 0

    command = [
        "powershell",
        "-NoProfile",
        "-ExecutionPolicy",
        "Bypass",
        "-File",
        str(helper_script),
        "-SourcePath",
        str(workbook_path),
        "-TargetPath",
        str(snapshot_path),
    ]
    last_message = "Excel could not create a readable snapshot."
    for attempt in range(4):
        result = subprocess.run(
            command,
            check=False,
            text=True,
            capture_output=True,
            creationflags=CREATE_NO_WINDOW,
            startupinfo=startupinfo,
        )
        if result.returncode == 0:
            return snapshot_path

        last_message = result.stderr.strip() or result.stdout.strip() or last_message
        snapshot_path.unlink(missing_ok=True)
        if "0x800AC472" not in last_message or attempt == 3:
            break
        time.sleep(1.5)

    raise RuntimeError(last_message)


def load_dashboard_workbook(workbook_path: Path) -> tuple[Any, Path | None]:
    if path_is_csv(workbook_path):
        snapshot_path = create_csv_snapshot(workbook_path)
        return load_workbook(snapshot_path, data_only=True), snapshot_path

    if os.name == "nt":
        try:
            snapshot_path = create_excel_snapshot(workbook_path)
            return load_workbook(snapshot_path, data_only=True), snapshot_path
        except Exception:
            pass

    try:
        return load_workbook(workbook_path, data_only=True), None
    except PermissionError:
        snapshot_path = create_excel_snapshot(workbook_path)
        return load_workbook(snapshot_path, data_only=True), snapshot_path


def cleanup_temp_file(path: Path | None) -> None:
    if path is None:
        return
    for _ in range(6):
        try:
            path.unlink(missing_ok=True)
            return
        except PermissionError:
            time.sleep(0.5)


def find_download_url(payload: Any) -> str | None:
    if isinstance(payload, dict):
        for key, value in payload.items():
            if "downloadurl" in key.lower() and isinstance(value, str) and value.startswith("http"):
                return value
            found = find_download_url(value)
            if found:
                return found
    if isinstance(payload, list):
        for item in payload:
            found = find_download_url(item)
            if found:
                return found
    return None


def extract_download_url_from_html(page_html: str) -> str | None:
    patterns = [
        r'"downloadUrl":"([^"]+)"',
        r'"@content\.downloadUrl":"([^"]+)"',
        r'"downloadUrl\\":\\"([^"]+)\\"',
    ]
    for pattern in patterns:
        match = re.search(pattern, page_html)
        if not match:
            continue
        candidate = match.group(1)
        candidate = candidate.replace("\\u0026", "&").replace("\\/", "/").replace('\\"', '"')
        return html.unescape(candidate)
    return None


def with_download_hint(url: str) -> str:
    parsed = urllib.parse.urlsplit(url)
    query = urllib.parse.parse_qsl(parsed.query, keep_blank_values=True)
    keys = {key.lower() for key, _ in query}
    if "download" not in keys:
        query.append(("download", "1"))
    return urllib.parse.urlunsplit((parsed.scheme, parsed.netloc, parsed.path, urllib.parse.urlencode(query), parsed.fragment))


def candidate_urls(url: str) -> list[str]:
    hinted = with_download_hint(url)
    return [url] if hinted == url else [url, hinted]


def ensure_excel_file(path: Path) -> None:
    if not zipfile.is_zipfile(path):
        raise ValueError(f"Downloaded file is not a valid Excel workbook: {path.name}")


def looks_like_excel_payload(payload: bytes) -> bool:
    return payload[:4] == b"PK\x03\x04"


def onedrive_badger_headers() -> dict[str, str]:
    token_payload = request_json(
        "https://api-badgerp.svc.ms/v1.0/token",
        headers={"Content-Type": "application/json", "User-Agent": "Mozilla/5.0"},
        data=json.dumps({"appId": "5cbed6ac-a083-4e14-b191-b4ba07653de2"}).encode("utf-8"),
    )
    token = token_payload.get("token")
    if not token:
        raise RuntimeError("Could not get OneDrive public access token.")
    return {
        "Authorization": f"Badger {token}",
        "Prefer": "autoredeem",
        "User-Agent": "Mozilla/5.0",
        "Accept": "application/json, */*",
    }


def download_onedrive_share(url: str) -> tuple[Path, str]:
    token = share_token(url)
    headers = onedrive_badger_headers()
    metadata = request_json(f"https://api.onedrive.com/v1.0/shares/{token}/driveItem", headers=headers)
    download_url = find_download_url(metadata)
    if download_url:
        data = request_bytes(download_url, headers={"User-Agent": "Mozilla/5.0"})
        filename = metadata.get("name") or Path(urllib.parse.urlsplit(download_url).path).name or "dashboard_source.xlsx"
        suffix = Path(filename).suffix or ".xlsx"
        target = write_temp_workbook(data, suffix)
        ensure_excel_file(target)
        return target, filename

    raise RuntimeError("OneDrive share metadata loaded, but no downloadable workbook URL was available.")


def download_workbook(url: str) -> tuple[Path, str]:
    headers = {"User-Agent": "Mozilla/5.0 Codex Dashboard Refresher"}
    last_error: Exception | None = None

    for candidate in candidate_urls(url):
        try:
            request = urllib.request.Request(candidate, headers=headers)
            with urllib.request.urlopen(request, timeout=60) as response:
                payload = response.read()
                final_url = response.geturl()
                content_type = response.headers.get("Content-Type", "")
                if looks_like_excel_payload(payload) or "spreadsheet" in content_type.lower():
                    suffix = Path(urllib.parse.urlsplit(final_url).path).suffix or ".xlsx"
                    target = write_temp_workbook(payload, suffix)
                    ensure_excel_file(target)
                    filename = Path(urllib.parse.urlsplit(final_url).path).name or "dashboard_source.xlsx"
                    return target, filename
                page_html = payload.decode("utf-8", errors="ignore")
                nested_url = extract_download_url_from_html(page_html)
                if nested_url:
                    nested_payload = request_bytes(nested_url, headers=headers)
                    suffix = Path(urllib.parse.urlsplit(nested_url).path).suffix or ".xlsx"
                    target = write_temp_workbook(nested_payload, suffix)
                    ensure_excel_file(target)
                    filename = Path(urllib.parse.urlsplit(nested_url).path).name or "dashboard_source.xlsx"
                    return target, filename
        except Exception as exc:  # noqa: BLE001
            last_error = exc

    if "onedrive" in url.lower() or "1drv.ms" in url.lower() or "sharepoint" in url.lower():
        return download_onedrive_share(url)

    raise RuntimeError(f"Could not download workbook from URL: {last_error}")


def find_default_workbook(bundle_dir: Path) -> Path | None:
    search_roots = candidate_source_dirs(bundle_dir)
    for root in search_roots:
        if not root.exists():
            continue
        for hint in WORKBOOK_NAME_HINTS:
            for pattern in (f"*{hint}*.xlsx", f"*{hint}*.xlsm"):
                preferred = sorted(root.glob(pattern))
                if preferred:
                    return preferred[0].resolve()
        for pattern in ("*.xlsx", "*.xlsm"):
            matches = sorted(root.glob(pattern))
            if matches:
                return matches[0].resolve()
    return None


def series_dataset(
    *,
    key: str,
    label: str,
    category: str,
    description: str,
    label_key: str,
    label_title: str,
    rows: list[dict[str, Any]],
    series: list[dict[str, Any]],
    facts: list[dict[str, str]],
    headline_value: str,
    headline_detail: str,
    tone: str,
    note: str,
    chart_kind: str = "line",
) -> dict[str, Any]:
    columns = [{"key": label_key, "label": label_title, "format": "text"}] + [
        {"key": item["key"], "label": item["label"], "format": item["format"]}
        for item in series
    ]
    chart_rows = [row for row in rows if any(row.get(item["key"]) is not None for item in series)]
    return {
        "key": key,
        "label": label,
        "category": category,
        "description": description,
        "headlineValue": headline_value,
        "headlineDetail": headline_detail,
        "tone": tone,
        "note": note,
        "facts": facts,
        "table": {
            "columns": columns,
            "rows": rows,
        },
        "chart": {
            "kind": chart_kind,
            "labels": [row[label_key] for row in chart_rows],
            "series": [
                {
                    "name": item["label"],
                    "key": item["key"],
                    "format": item["format"],
                    "color": item["color"],
                    "values": [row.get(item["key"]) for row in chart_rows],
                    "style": item.get("style", "solid"),
                    "showDots": item.get("showDots", True),
                    "strokeWidth": item.get("strokeWidth", 3),
                }
                for item in series
            ],
        },
    }


def parse_housekeeping(ws) -> dict[str, Any]:
    rows = []
    average_92m = to_float(ws["B3"].value)
    average_12m = to_float(ws["C3"].value)
    rows.append({"period": "2025 Avg", "score92m": average_92m, "score12m": average_12m})
    for row_num in range(4, LIVE_DATA_MAX_ROW + 1):
        week = to_int(ws[f"A{row_num}"].value)
        score_92m = to_float(ws[f"B{row_num}"].value)
        score_12m = to_float(ws[f"C{row_num}"].value)
        if week is None and score_92m is None and score_12m is None:
            continue
        rows.append({"period": f"Week {week}", "score92m": score_92m, "score12m": score_12m})

    weekly_rows = [row for row in rows if row["period"].startswith("Week")]
    latest = latest_non_null(weekly_rows, "score92m")
    best = max_row(weekly_rows, "score92m")
    dataset = series_dataset(
        key="housekeeping",
        label="Housekeeping",
        category="Quality",
        description="Weekly housekeeping performance with both the 92M and 12M score streams.",
        label_key="period",
        label_title="Period",
        rows=rows,
        series=[
            {"key": "score92m", "label": "92M", "format": "percent", "color": "#00cfff"},
            {"key": "score12m", "label": "12M", "format": "percent", "color": "#ffb703"},
        ],
        facts=[
            {"label": "Latest", "value": format_percent(latest["score92m"], 0) if latest else "-"},
            {"label": "Target", "value": format_percent(0.80, 0)},
            {"label": "Average", "value": format_percent(average_92m, 0)},
            {"label": "Best Week", "value": f"{best['period']} - {format_percent(best['score92m'], 0)}" if best else "-"},
        ],
        headline_value=format_percent(latest["score92m"], 0) if latest else "-",
        headline_detail=f"{latest['period']} 92M reading" if latest else "No reading yet",
        tone=tone_housekeeping(latest["score92m"] if latest else None),
        note="Shows weekly housekeeping performance against the green 80% target line and the 2025 average benchmark.",
    )
    add_target_line(dataset, value=0.80, label="92M Target", format_type="percent")
    add_trendline(dataset, source_key="score92m", label="92M Trend")
    return dataset


def parse_single_series_block(
    ws,
    *,
    key: str,
    label: str,
    category: str,
    description: str,
    label_key: str,
    label_title: str,
    label_prefix: str,
    label_col: str,
    value_col: str,
    start_row: int,
    end_row: int,
    series_key: str,
    series_label: str,
    series_format: str,
    color: str,
    target_value: float,
    tone_resolver: Callable[[float | None], str],
    higher_is_better: bool = True,
    note: str = "",
) -> dict[str, Any]:
    rows = []
    for row_num in range(start_row, end_row + 1):
        raw_label = to_int(ws[f"{label_col}{row_num}"].value)
        raw_value = to_float(ws[f"{value_col}{row_num}"].value)
        if raw_label is None and raw_value is None:
            continue
        rows.append({label_key: f"{label_prefix} {raw_label}", series_key: raw_value})

    latest = latest_non_null(rows, series_key)
    best = max_row(rows, series_key) if higher_is_better else min_row(rows, series_key)
    values = [row[series_key] for row in rows if row.get(series_key) is not None]
    average_value = mean(values)
    dataset = series_dataset(
        key=key,
        label=label,
        category=category,
        description=description,
        label_key=label_key,
        label_title=label_title,
        rows=rows,
        series=[{"key": series_key, "label": series_label, "format": series_format, "color": color}],
        facts=[
            {"label": "Latest", "value": format_percent(latest[series_key], 1) if latest else "-"},
            {"label": "Target", "value": format_percent(target_value, 1)},
            {"label": "Average", "value": format_percent(average_value, 1)},
            {"label": "Best", "value": f"{best[label_key]} - {format_percent(best[series_key], 1)}" if best else "-"},
            {"label": "Readings", "value": str(len(values))},
        ],
        headline_value=format_percent(latest[series_key], 1) if latest else "-",
        headline_detail=f"{latest[label_key]} reading" if latest else "No reading yet",
        tone=tone_resolver(latest[series_key] if latest else None),
        note=note,
    )
    add_target_line(dataset, value=target_value, label="Target", format_type=series_format)
    add_trendline(dataset, source_key=series_key, label="Trend")
    return dataset


def parse_sku_share(ws) -> dict[str, Any]:
    rows = []
    for row_num in range(3, LIVE_DATA_MAX_ROW + 1):
        picker = clean_text(ws[f"T{row_num}"].value)
        count = to_float(ws[f"U{row_num}"].value)
        share = to_float(ws[f"V{row_num}"].value)
        if not picker:
            continue
        rows.append({"picker": picker, "count": count, "share": share})

    top_picker = max_row(rows, "count")
    total_units = sum(row["count"] for row in rows if row.get("count") is not None)
    return {
        "key": "sku-share",
        "label": "Picker Contribution",
        "category": "Ranking",
        "description": "Executive view of picked-SKU volume concentration across the active picker roster.",
        "headlineValue": format_number(top_picker["count"], 0) if top_picker else "-",
        "headlineDetail": f"{top_picker['picker']} leads picked volume" if top_picker else "No picker data yet",
        "tone": "good",
        "note": "Donut view shows contribution mix while the table keeps exact picked counts and shares visible.",
        "facts": [
            {"label": "Leading Picker", "value": f"{top_picker['picker']} - {format_number(top_picker['count'], 0)}" if top_picker else "-"},
            {"label": "Lead Share", "value": format_percent(top_picker["share"], 1) if top_picker else "-"},
            {"label": "Total SKUs Picked", "value": format_number(total_units, 0)},
            {"label": "Active Pickers", "value": str(len(rows))},
        ],
        "table": {
            "columns": [
                {"key": "picker", "label": "Picker", "format": "text"},
                {"key": "count", "label": "Picked SKUs", "format": "integer"},
                {"key": "share", "label": "Share", "format": "percent"},
            ],
            "rows": rows,
        },
        "chart": {
            "kind": "donut",
            "series": [
                {"name": row["picker"], "value": row["count"], "share": row["share"], "color": color}
                for row, color in zip(
                    rows,
                    ["#ff8a5b", "#26d0ce", "#ffc857", "#8ddf6e", "#8a7dff"],
                    strict=False,
                )
            ],
        },
    }


def find_monthly_sku_layout(ws) -> dict[str, Any]:
    for col_num in range(2, ws.max_column + 1):
        current = clean_text(ws.cell(row=2, column=col_num).value)
        next_value = clean_text(ws.cell(row=2, column=col_num + 1).value)
        if current != "2024" or next_value != "2025":
            continue

        month_col = col_num - 1
        year_2026_cols: list[int] = []
        cursor = col_num + 2
        while cursor <= ws.max_column:
            header = clean_text(ws.cell(row=2, column=cursor).value)
            if header.startswith("2026"):
                year_2026_cols.append(cursor)
                cursor += 1
                continue
            break

        if year_2026_cols:
            return {
                "month_col": month_col,
                "year_2024_col": col_num,
                "year_2025_col": col_num + 1,
                "year_2026_cols": year_2026_cols,
            }

    raise ValueError("Could not locate the monthly SKU section in the DATA sheet.")


def find_liseo_layout(ws) -> dict[str, int]:
    section_col = None
    for col_num in range(1, ws.max_column + 1):
        if clean_text(ws.cell(row=1, column=col_num).value) == "Liseo Assembly vs Backorders":
            section_col = col_num
            break

    if section_col is None:
        raise ValueError("Could not locate the Liseo assembly section in the DATA sheet.")

    columns: dict[str, int] = {}
    for label in ("Month", "Assembled", "Backorders", "Fill Rate"):
        for col_num in range(section_col, min(ws.max_column, section_col + 6) + 1):
            if clean_text(ws.cell(row=2, column=col_num).value) == label:
                columns[label] = col_num
                break
        if label not in columns:
            raise ValueError(f"Could not find '{label}' inside the Liseo assembly section.")
    return columns


def parse_monthly_sku(ws) -> dict[str, Any]:
    layout = find_monthly_sku_layout(ws)

    def year_2026_key(label: str, index: int) -> str:
        suffix = clean_text(label).replace("2026", "", 1).strip()
        if not suffix:
            return "y2026" if index == 1 else f"y2026_{index}"
        parts = [part for part in re.split(r"[^0-9a-zA-Z]+", suffix) if part]
        if not parts:
            return "y2026" if index == 1 else f"y2026_{index}"
        return "y2026" + "".join(part[:1].upper() + part[1:].lower() for part in parts)

    year_2026_specs = []
    for index, col_num in enumerate(layout["year_2026_cols"], start=1):
        header = clean_text(ws.cell(row=2, column=col_num).value) or f"2026 {index}"
        year_2026_specs.append(
            {
                "column": col_num,
                "key": year_2026_key(header, index),
                "label": header,
            }
        )

    rows = []
    for row_num in range(3, LIVE_DATA_MAX_ROW + 1):
        month = format_month_label(ws.cell(row=row_num, column=layout["month_col"]).value)
        if not month:
            continue
        row_data = {
            "month": month,
            "y2024": to_float(ws.cell(row=row_num, column=layout["year_2024_col"]).value),
            "y2025": to_float(ws.cell(row=row_num, column=layout["year_2025_col"]).value),
        }
        for spec in year_2026_specs:
            row_data[spec["key"]] = to_float(ws.cell(row=row_num, column=spec["column"]).value)
        if len(year_2026_specs) > 1:
            total_value = sum(value for value in (row_data.get(spec["key"]) for spec in year_2026_specs) if value is not None)
            row_data["y2026Total"] = total_value if total_value else (0 if any(row_data.get(spec["key"]) == 0 for spec in year_2026_specs) else None)
        rows.append(row_data)

    primary_2026_key = "y2026Total" if len(year_2026_specs) > 1 else year_2026_specs[0]["key"]
    george_spec = next((spec for spec in year_2026_specs if "george" in spec["label"].lower()), None)
    cpt_spec = next((spec for spec in year_2026_specs if "cpt" in spec["label"].lower()), None)
    ytd_2026 = sum(row[primary_2026_key] for row in rows if row.get(primary_2026_key) is not None)
    live_rows = [row for row in rows if row.get(primary_2026_key) is not None]
    ytd_2024 = sum(row["y2024"] for row in live_rows if row.get("y2024") is not None)
    ytd_2025 = sum(row["y2025"] for row in live_rows if row.get("y2025") is not None)
    ytd_values = sorted([ytd_2024, ytd_2025, ytd_2026], reverse=True)
    ytd_rank = ytd_values.index(ytd_2026) + 1 if ytd_values else None
    best_2025 = max_row(rows, "y2025")
    latest_2026 = latest_non_null(rows, primary_2026_key)
    ytd_2026_george = 0 if george_spec else None
    if george_spec:
        ytd_2026_george = sum(row[george_spec["key"]] for row in rows if row.get(george_spec["key"]) is not None)
    table_columns = [
        {"key": "month", "label": "Month", "format": "text"},
        {"key": "y2024", "label": "2024", "format": "integer"},
        {"key": "y2025", "label": "2025", "format": "integer"},
    ] + [{"key": spec["key"], "label": spec["label"], "format": "integer"} for spec in year_2026_specs]
    if len(year_2026_specs) > 1:
        table_columns.append({"key": "y2026Total", "label": "SKUs Picked", "format": "integer"})

    chart_series = [
        {"name": "2024", "key": "y2024", "format": "integer", "color": "#3b82f6", "values": [row["y2024"] for row in rows], "style": "solid", "showDots": True, "strokeWidth": 2},
        {"name": "2025 Benchmark", "key": "y2025", "format": "integer", "color": "#39ff88", "values": [row["y2025"] for row in rows], "style": "dashed", "showDots": False, "strokeWidth": 2},
    ]
    if len(year_2026_specs) > 1:
        chart_series.append(
            {
                "name": "SKUs Picked",
                "key": "y2026Total",
                "format": "integer",
                "color": "#00cfff",
                "values": [row.get("y2026Total") for row in rows],
                "style": "solid",
                "showDots": True,
                "strokeWidth": 4,
            }
        )
        ordered_components = []
        if george_spec:
            ordered_components.append((george_spec, "#8b5cf6", 3))
        if cpt_spec and cpt_spec != george_spec:
            ordered_components.append((cpt_spec, "#ff8a5b", 2.5))
        ordered_component_keys = {component[0]["key"] for component in ordered_components}
        for spec in year_2026_specs:
            if spec["key"] in ordered_component_keys:
                continue
            ordered_components.append((spec, "#ffd54f", 2.5))
        for spec, color, stroke_width in ordered_components:
            chart_series.append(
                {
                    "name": spec["label"],
                    "key": spec["key"],
                    "format": "integer",
                    "color": color,
                    "values": [row.get(spec["key"]) for row in rows],
                    "style": "solid",
                    "showDots": True,
                    "strokeWidth": stroke_width,
                }
            )
    else:
        spec = year_2026_specs[0]
        chart_series.append(
            {
                "name": "2026 Actual",
                "key": spec["key"],
                "format": "integer",
                "color": "#00cfff",
                "values": [row.get(spec["key"]) for row in rows],
                "style": "solid",
                "showDots": True,
                "strokeWidth": 4,
            }
        )
    chart_series.append(
        {
            "name": "2026 Run-Rate Trend",
            "format": "integer",
            "color": "#f8fafc",
            "values": build_trendline([row.get(primary_2026_key) for row in rows]),
            "style": "dotted",
            "showDots": False,
            "strokeWidth": 2,
        }
    )
    return {
        "key": "sku-monthly",
        "label": "SKUs Picked Trend",
        "category": "Volume",
        "description": "Executive monthly view of picked-SKU volume versus prior years, with the combined current-year total shown alongside the CPT and George split whenever that breakdown exists.",
        "headlineValue": format_number(ytd_2026, 0),
        "headlineDetail": "2026 YTD SKUs picked across CPT and George" if len(year_2026_specs) > 1 and george_spec else "2026 YTD SKUs picked",
        "tone": tone_from_rank(ytd_rank),
        "note": "Table view keeps the combined SKUs Picked total visible alongside the 2026 CPT and George split, while chart view reads that combined total against the 2025 benchmark.",
        "facts": [
            {"label": "2026 SKUs Picked YTD" if len(year_2026_specs) > 1 else "2026 YTD", "value": format_number(ytd_2026, 0)},
            *([{"label": "George Contribution YTD", "value": format_number(ytd_2026_george, 0)}] if george_spec else []),
            {"label": "Latest Month Total", "value": f"{latest_2026['month']} - {format_number(latest_2026[primary_2026_key], 0)}" if latest_2026 else "-"},
            {"label": "Best 2025 Month", "value": f"{best_2025['month']} - {format_number(best_2025['y2025'], 0)}" if best_2025 else "-"},
            {"label": "YTD Position", "value": f"{ytd_rank} of 3" if ytd_rank else "-"},
        ],
        "table": {
            "columns": table_columns,
            "rows": rows,
        },
        "chart": {
            "kind": "line",
            "labels": [row["month"] for row in rows],
            "series": chart_series,
        },
    }


def parse_assembly_backorders(ws) -> dict[str, Any]:
    layout = find_liseo_layout(ws)
    rows = []
    for row_num in range(3, LIVE_DATA_MAX_ROW + 1):
        month = format_month_label(ws.cell(row=row_num, column=layout["Month"]).value)
        if not month:
            continue
        assembled = to_float(ws.cell(row=row_num, column=layout["Assembled"]).value)
        backorders = to_float(ws.cell(row=row_num, column=layout["Backorders"]).value)
        fill_rate = to_float(ws.cell(row=row_num, column=layout["Fill Rate"]).value)
        if fill_rate is None and assembled not in (None, 0) and backorders is not None:
            fill_rate = 1 - (backorders / assembled)
        rows.append(
            {
                "month": month,
                "assembled": assembled,
                "backorders": backorders,
                "fillRate": fill_rate,
            }
        )

    latest = latest_non_null(rows, "fillRate")
    best = max_row(rows, "fillRate")
    average_value = mean([row["fillRate"] for row in rows if row.get("fillRate") is not None])
    dataset = {
        "key": "assembly-backorders",
        "label": "Assembly vs Backorders",
        "category": "Production",
        "description": "Monthly Liseo assembly output against backorders, with fill rate calculated from the live workbook.",
        "headlineValue": format_percent(latest["fillRate"], 1) if latest else "-",
        "headlineDetail": f"{latest['month']} fill rate" if latest else "No reading yet",
        "tone": tone_assembly_fill_rate(latest["fillRate"] if latest else None),
        "note": "Table view keeps the assembled and backorder counts, while chart view focuses on the monthly fill-rate line against the green 85% target.",
        "facts": [
            {"label": "Latest", "value": format_percent(latest["fillRate"], 1) if latest else "-"},
            {"label": "Target", "value": format_percent(0.85, 0)},
            {"label": "Average", "value": format_percent(average_value, 1)},
            {"label": "Best", "value": f"{best['month']} - {format_percent(best['fillRate'], 1)}" if best else "-"},
        ],
        "table": {
            "columns": [
                {"key": "month", "label": "Month", "format": "text"},
                {"key": "assembled", "label": "Assembled", "format": "integer"},
                {"key": "backorders", "label": "Backorders", "format": "integer"},
                {"key": "fillRate", "label": "Fill Rate", "format": "percent"},
            ],
            "rows": rows,
        },
        "chart": {
            "kind": "line",
            "labels": [row["month"] for row in rows if row.get("fillRate") is not None],
            "series": [
                {
                    "name": "Fill Rate",
                    "key": "fillRate",
                    "format": "percent",
                    "color": "#8b5cf6",
                    "values": [row["fillRate"] for row in rows if row.get("fillRate") is not None],
                    "style": "solid",
                    "showDots": True,
                    "strokeWidth": 3,
                }
            ],
        },
    }
    add_target_line(dataset, value=0.85, label="Target", format_type="percent")
    add_trendline(dataset, source_key="fillRate", label="Trend")
    return dataset


def parse_points_yoy(ws) -> dict[str, Any]:
    rows = []
    for row_num in range(19, 31):
        month = format_month_label(ws[f"A{row_num}"].value)
        if not month:
            continue
        rows.append(
            {
                "month": month,
                "y2021": to_float(ws[f"B{row_num}"].value),
                "y2022": to_float(ws[f"C{row_num}"].value),
                "y2023": to_float(ws[f"D{row_num}"].value),
                "y2024": to_float(ws[f"E{row_num}"].value),
                "y2025": to_float(ws[f"F{row_num}"].value),
                "y2026": to_float(ws[f"G{row_num}"].value),
                "total": to_float(ws[f"H{row_num}"].value),
            }
        )

    total_2026 = to_float(ws["G31"].value)
    total_2025 = to_float(ws["F31"].value)
    peak_month = None
    for year_key in ("y2021", "y2022", "y2023", "y2024", "y2025", "y2026"):
        candidate = max_row(rows, year_key)
        if not candidate:
            continue
        if peak_month is None or candidate[year_key] > peak_month["value"]:
            peak_month = {"month": candidate["month"], "year": year_key[1:], "value": candidate[year_key]}

    return {
        "key": "points-yoy",
        "label": "Liseo Points YOY",
        "category": "Year over Year",
        "description": "Monthly points history across six years with a total-per-month column kept in the table view.",
        "headlineValue": format_number(total_2026, 1),
        "headlineDetail": "2026 year-to-date points",
        "tone": "good",
        "note": "The line view exposes long-term seasonality and where 2026 is still building up.",
        "facts": [
            {"label": "2026 YTD", "value": format_number(total_2026, 1)},
            {"label": "2025 Total", "value": format_number(total_2025, 0)},
            {"label": "Peak Month", "value": f"{peak_month['month']} {peak_month['year']} - {format_number(peak_month['value'], 1)}" if peak_month else "-"},
            {"label": "Benchmark", "value": "2025 line"},
        ],
        "table": {
            "columns": [
                {"key": "month", "label": "Month", "format": "text"},
                {"key": "y2021", "label": "2021", "format": "decimal1"},
                {"key": "y2022", "label": "2022", "format": "decimal1"},
                {"key": "y2023", "label": "2023", "format": "decimal1"},
                {"key": "y2024", "label": "2024", "format": "decimal1"},
                {"key": "y2025", "label": "2025", "format": "decimal1"},
                {"key": "y2026", "label": "2026", "format": "decimal1"},
                {"key": "total", "label": "Total", "format": "decimal1"},
            ],
            "rows": rows + [
                {
                    "month": "Total",
                    "y2021": to_float(ws["B31"].value),
                    "y2022": to_float(ws["C31"].value),
                    "y2023": to_float(ws["D31"].value),
                    "y2024": to_float(ws["E31"].value),
                    "y2025": to_float(ws["F31"].value),
                    "y2026": total_2026,
                    "total": to_float(ws["H31"].value),
                }
            ],
        },
        "chart": {
            "kind": "line",
            "labels": [row["month"] for row in rows],
            "series": [
                {"name": "2021", "key": "y2021", "format": "decimal1", "color": "#6c7385", "values": [row["y2021"] for row in rows], "style": "solid", "showDots": False, "strokeWidth": 1.5},
                {"name": "2022", "key": "y2022", "format": "decimal1", "color": "#ff8a5b", "values": [row["y2022"] for row in rows], "style": "solid", "showDots": False, "strokeWidth": 1.5},
                {"name": "2023", "key": "y2023", "format": "decimal1", "color": "#ffd54f", "values": [row["y2023"] for row in rows], "style": "solid", "showDots": False, "strokeWidth": 1.5},
                {"name": "2024", "key": "y2024", "format": "decimal1", "color": "#7ac7b1", "values": [row["y2024"] for row in rows], "style": "solid", "showDots": False, "strokeWidth": 1.5},
                {"name": "2025 Target", "key": "y2025", "format": "decimal1", "color": "#00cfff", "values": [row["y2025"] for row in rows], "style": "dashed", "showDots": False, "strokeWidth": 2},
                {"name": "2026 Actual", "key": "y2026", "format": "decimal1", "color": "#ff6b74", "values": [row["y2026"] for row in rows], "style": "solid", "showDots": True, "strokeWidth": 3.5},
                {"name": "2026 Trend", "format": "decimal1", "color": "#f5efeb", "values": build_trendline([row["y2026"] for row in rows]), "style": "dotted", "showDots": False, "strokeWidth": 2},
            ],
        },
    }


def build_dashboard(workbook_path: Path) -> dict[str, Any]:
    workbook = None
    snapshot_path: Path | None = None
    try:
        workbook, snapshot_path = load_dashboard_workbook(workbook_path)
        ws = workbook["DATA"]
        source_mtime = dt.datetime.fromtimestamp(workbook_path.stat().st_mtime, dt.timezone.utc)
        generated_at = dt.datetime.now(dt.timezone.utc)

        housekeeping = parse_housekeeping(ws)
        jhb = parse_single_series_block(
            ws,
            key="container-jhb",
            label="Container Accuracy JHB",
            category="Accuracy",
            description="Load-by-load accuracy readings for the Johannesburg container stream.",
            label_key="load",
            label_title="Load",
            label_prefix="Load",
            label_col="F",
            value_col="G",
            start_row=3,
            end_row=LIVE_DATA_MAX_ROW,
            series_key="accuracy",
            series_label="Accuracy",
            series_format="percent",
            color="#00cfff",
            target_value=0.99,
            tone_resolver=tone_accuracy,
            note="Higher is better. Missing loads are left blank so the chart stays honest.",
        )
        george = parse_single_series_block(
            ws,
            key="container-george",
            label="Container Accuracy George",
            category="Accuracy",
            description="Load-by-load accuracy readings for the George container stream.",
            label_key="load",
            label_title="Load",
            label_prefix="Load",
            label_col="I",
            value_col="J",
            start_row=3,
            end_row=LIVE_DATA_MAX_ROW,
            series_key="accuracy",
            series_label="Accuracy",
            series_format="percent",
            color="#3b82f6",
            target_value=0.99,
            tone_resolver=tone_accuracy,
            note="A shorter series today, but the same live refresh loop will grow it automatically.",
        )
        urgent = parse_single_series_block(
            ws,
            key="urgent-orders",
            label="Wholesaler Urgent Orders",
            category="Exceptions",
            description="Weekly urgent-order rate. Lower is healthier here.",
            label_key="week",
            label_title="Week",
            label_prefix="Week",
            label_col="M",
            value_col="N",
            start_row=3,
            end_row=LIVE_DATA_MAX_ROW,
            series_key="rate",
            series_label="Urgent Rate",
            series_format="percent",
            color="#ffb703",
            target_value=0.05,
            tone_resolver=tone_urgent_orders,
            higher_is_better=False,
            note="This one flips the interpretation: the closer to zero, the better.",
        )
        dispatch = parse_single_series_block(
            ws,
            key="dispatch-accuracy",
            label="Dispatch Accuracy",
            category="Accuracy",
            description="Weekly dispatch accuracy performance.",
            label_key="week",
            label_title="Week",
            label_prefix="Week",
            label_col="Q",
            value_col="R",
            start_row=3,
            end_row=LIVE_DATA_MAX_ROW,
            series_key="accuracy",
            series_label="Accuracy",
            series_format="percent",
            color="#ff5d73",
            target_value=0.99,
            tone_resolver=tone_accuracy,
            note="A near-perfect trend, so the chart leans on fine-grained percentage labels.",
        )
        sku_share = parse_sku_share(ws)
        monthly_sku = parse_monthly_sku(ws)
        assemblies = parse_assembly_backorders(ws)
        return {
            "title": "Operations Live Dashboard",
            "subtitle": "PPT Presentation Source",
            "sourceName": workbook_path.name,
            "generatedAt": generated_at.isoformat(),
            "sourceModifiedAt": source_mtime.isoformat(),
            "refreshSeconds": 60,
            "summaryCards": [
                {
                    "label": "Housekeeping",
                    "value": housekeeping["headlineValue"],
                    "detail": housekeeping["headlineDetail"],
                    "tone": housekeeping["tone"],
                },
                {
                    "label": "JHB Accuracy",
                    "value": jhb["headlineValue"],
                    "detail": jhb["headlineDetail"],
                    "tone": jhb["tone"],
                },
                {
                    "label": "George Accuracy",
                    "value": george["headlineValue"],
                    "detail": george["headlineDetail"],
                    "tone": george["tone"],
                },
                {
                    "label": "Dispatch Accuracy",
                    "value": dispatch["headlineValue"],
                    "detail": dispatch["headlineDetail"],
                    "tone": dispatch["tone"],
                },
                {
                    "label": "Urgent Orders",
                    "value": urgent["headlineValue"],
                    "detail": urgent["headlineDetail"],
                    "tone": urgent["tone"],
                },
                {
                    "label": "Top Picker Volume",
                    "value": sku_share["headlineValue"],
                    "detail": sku_share["headlineDetail"],
                    "tone": sku_share["tone"],
                },
                {
                    "label": "SKUs Picked YTD",
                    "value": monthly_sku["headlineValue"],
                    "detail": monthly_sku["headlineDetail"],
                    "tone": monthly_sku["tone"],
                },
                {
                    "label": "Assembly vs BO",
                    "value": assemblies["headlineValue"],
                    "detail": assemblies["headlineDetail"],
                    "tone": assemblies["tone"],
                },
            ],
            "datasets": [
                housekeeping,
                jhb,
                george,
                urgent,
                dispatch,
                sku_share,
                monthly_sku,
                assemblies,
            ],
        }
    finally:
        if workbook is not None:
            workbook.close()
        if snapshot_path is not None:
            cleanup_temp_file(snapshot_path)


def resolve_workbook(args: argparse.Namespace, bundle_dir: Path) -> tuple[Path, str]:
    if args.workbook:
        path = Path(args.workbook).expanduser().resolve()
        if not path.exists():
            raise FileNotFoundError(f"Workbook not found: {path}")
        return path, path.name

    if args.workbook_url:
        temp_path, filename = download_workbook(args.workbook_url)
        return temp_path, filename

    path = choose_preferred_source(None, bundle_dir)
    if path is None:
        raise FileNotFoundError("Could not find a local workbook or CSV export to build the dashboard from.")
    return path.resolve(), path.name


def refresh_dashboard_data(*, workbook: str | None = None, workbook_url: str | None = None, output: str | Path = DEFAULT_OUTPUT) -> Path:
    script_dir = Path(__file__).resolve().parent
    bundle_dir = script_dir.parent
    args = argparse.Namespace(workbook=workbook, workbook_url=workbook_url, output=output)
    workbook_path, _ = resolve_workbook(args, bundle_dir)

    output_path = Path(output).expanduser()
    if not output_path.is_absolute():
        output_path = (bundle_dir / output_path).resolve()

    payload = build_dashboard(workbook_path)
    output_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    return output_path


def main() -> None:
    args = parse_args()
    refresh_dashboard_data(
        workbook=args.workbook,
        workbook_url=args.workbook_url or os.environ.get("WORKBOOK_URL"),
        output=args.output,
    )


if __name__ == "__main__":
    main()
