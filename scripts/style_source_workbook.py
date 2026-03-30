from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


THIN = Side(style="thin", color="27404D")
MEDIUM = Side(style="medium", color="375869")
WHITE = "F2F7F9"
MUTED = "D7E5EA"
HEADER_FILL = PatternFill("solid", fgColor="103040")
SUBHEAD_FILL = PatternFill("solid", fgColor="173B4D")
SECTION_FILL = PatternFill("solid", fgColor="0F2433")
ACCENT_TEAL = PatternFill("solid", fgColor="1D7C78")
ACCENT_BLUE = PatternFill("solid", fgColor="246BA0")
ACCENT_AMBER = PatternFill("solid", fgColor="A46A1D")
ACCENT_CORAL = PatternFill("solid", fgColor="A04B36")
ACCENT_VIOLET = PatternFill("solid", fgColor="6652A5")
LIGHT_ROW = PatternFill("solid", fgColor="EFF7FA")
ALT_ROW = PatternFill("solid", fgColor="E6F1F6")
TOTAL_FILL = PatternFill("solid", fgColor="D9E8EF")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Apply a visual cleanup pass to the PPT presentation source workbook.")
    parser.add_argument("--workbook", required=True, help="Workbook to style.")
    parser.add_argument("--output", help="Optional output path. Defaults to updating the workbook in place.")
    return parser.parse_args()


def apply_title(cell, fill):
    cell.fill = fill
    cell.font = Font(name="Aptos Display", size=12, bold=True, color=WHITE)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(left=MEDIUM, right=MEDIUM, top=MEDIUM, bottom=MEDIUM)


def apply_header(cell):
    cell.fill = SUBHEAD_FILL
    cell.font = Font(name="Aptos", size=10, bold=True, color=WHITE)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def apply_data(cell, *, align="center", fill=None, number_format=None, bold=False):
    cell.fill = fill or LIGHT_ROW
    cell.font = Font(name="Aptos", size=10, bold=bold, color="10212A")
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    if number_format:
        cell.number_format = number_format


def row_fill(row_num: int) -> PatternFill:
    return ALT_ROW if row_num % 2 else LIGHT_ROW


def style_workbook(path: Path, output_path: Path) -> None:
    wb = load_workbook(path)
    ws = wb["DATA"]

    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "1B6F7A"
    ws.freeze_panes = "A3"
    ws.sheet_view.zoomScale = 92

    widths = {
        "A": 14, "B": 11, "C": 11, "D": 4, "E": 4,
        "F": 11, "G": 13, "H": 4, "I": 11, "J": 13, "K": 4,
        "L": 14, "M": 10, "N": 12, "O": 4, "P": 6, "Q": 10, "R": 12,
        "S": 4, "T": 14, "U": 11, "V": 11, "W": 4, "X": 10, "Y": 11, "Z": 11, "AA": 11,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    for row_num, height in {
        1: 24, 2: 22, 3: 22, 16: 8, 17: 24, 18: 22, 31: 22,
    }.items():
        ws.row_dimensions[row_num].height = height

    for merged, fill in {
        "A1:D1": ACCENT_TEAL,
        "F1:G1": ACCENT_BLUE,
        "I1:J1": ACCENT_BLUE,
        "L1:N1": ACCENT_AMBER,
        "P1:R1": ACCENT_TEAL,
        "T1:V1": ACCENT_CORAL,
        "X1:AA1": ACCENT_VIOLET,
        "A17:H17": HEADER_FILL,
    }.items():
        top_left = ws[merged.split(":")[0]]
        apply_title(top_left, fill)
        for row in ws[merged]:
            for cell in row:
                cell.border = Border(left=MEDIUM, right=MEDIUM, top=MEDIUM, bottom=MEDIUM)

    for row in ws["A2:AA2"][0]:
        if row.coordinate in {"D2", "E2", "H2", "K2", "O2", "R2", "S2", "W2"}:
            row.fill = PatternFill("solid", fgColor="0C1B26")
            continue
        apply_header(row)

    for row in ws["A18:H18"][0]:
        apply_header(row)

    for row in ws["A3:A15"]:
        for cell in row:
            apply_data(cell, align="left" if cell.row == 3 else "center", fill=row_fill(cell.row), bold=cell.row == 3)
    for row in ws["B3:C15"]:
        for cell in row:
            apply_data(cell, fill=row_fill(cell.row), number_format="0%")

    for row in ws["F3:F15"]:
        for cell in row:
            apply_data(cell, fill=row_fill(cell.row))
    for row in ws["G3:G15"]:
        for cell in row:
            apply_data(cell, fill=row_fill(cell.row), number_format="0.0%")

    for row in ws["I3:I15"]:
        for cell in row:
            apply_data(cell, fill=row_fill(cell.row))
    for row in ws["J3:J15"]:
        for cell in row:
            apply_data(cell, fill=row_fill(cell.row), number_format="0.0%")

    for row in ws["M3:M15"]:
        for cell in row:
            apply_data(cell, fill=row_fill(cell.row))
    for row in ws["N3:N15"]:
        for cell in row:
            apply_data(cell, fill=row_fill(cell.row), number_format="0.0%")

    for row in ws["Q3:Q15"]:
        for cell in row:
            apply_data(cell, fill=row_fill(cell.row))
    for row in ws["R3:R15"]:
        for cell in row:
            apply_data(cell, fill=row_fill(cell.row), number_format="0.0%")

    for row in ws["T3:T7"]:
        for cell in row:
            apply_data(cell, align="left", fill=row_fill(cell.row))
    for row in ws["U3:U7"]:
        for cell in row:
            apply_data(cell, fill=row_fill(cell.row), number_format="#,##0")
    for row in ws["V3:V7"]:
        for cell in row:
            apply_data(cell, fill=row_fill(cell.row), number_format="0.0%")

    for row in ws["X3:X14"]:
        for cell in row:
            apply_data(cell, align="left", fill=row_fill(cell.row))
    for row in ws["Y3:AA14"]:
        for cell in row:
            apply_data(cell, fill=row_fill(cell.row), number_format="#,##0")

    for row in ws["A19:A31"]:
        for cell in row:
            apply_data(cell, align="left", fill=row_fill(cell.row), bold=cell.row == 31)
    for row in ws["B19:H31"]:
        for cell in row:
            apply_data(cell, fill=row_fill(cell.row), number_format="0.0")

    for cell in ws["A3:C3"][0]:
        apply_data(cell, fill=TOTAL_FILL, bold=True, number_format="0%")
    for cell in ws["A31:H31"][0]:
        apply_data(cell, fill=TOTAL_FILL, bold=True, number_format="0.0" if cell.column > 1 else None)

    ws["A3"].fill = HEADER_FILL
    ws["A3"].font = Font(name="Aptos", size=10, bold=True, color=WHITE)
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A31"].fill = HEADER_FILL
    ws["A31"].font = Font(name="Aptos", size=10, bold=True, color=WHITE)
    ws["A31"].alignment = Alignment(horizontal="left", vertical="center")

    for range_string in ["B4:C14", "G3:G15", "J3:J15", "N3:N15", "Q3:Q15", "V3:V7"]:
        ws.conditional_formatting.add(
            range_string,
            ColorScaleRule(
                start_type="num", start_value=0, start_color="F26C4F",
                mid_type="num", mid_value=0.75, mid_color="F7D266",
                end_type="num", end_value=1, end_color="49C6A7",
            ),
        )

    for range_string in ["U3:U7", "Y3:AA14", "B19:H31"]:
        ws.conditional_formatting.add(
            range_string,
            ColorScaleRule(
                start_type="min", start_color="DCEFF4",
                mid_type="percentile", mid_value=50, mid_color="89C7D6",
                end_type="max", end_color="1A7A8C",
            ),
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> None:
    args = parse_args()
    workbook_path = Path(args.workbook).expanduser().resolve()
    output_path = Path(args.output).expanduser().resolve() if args.output else workbook_path
    style_workbook(workbook_path, output_path)


if __name__ == "__main__":
    main()
